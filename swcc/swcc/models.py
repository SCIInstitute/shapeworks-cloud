from __future__ import annotations

import itertools
import logging
from pathlib import Path, PurePath
import re
from tempfile import TemporaryDirectory

try:
    from typing import (
        Any,
        Dict,
        Generic,
        Iterator,
        Literal,
        Optional,
        Tuple,
        Type,
        TypeVar,
        Union,
        get_args,
    )
except ImportError:
    from typing import (
        Any,
        Dict,
        Generic,
        Iterator,
        Optional,
        Tuple,
        Type,
        TypeVar,
        Union,
    )
    from typing_extensions import (  # type: ignore
        Literal,
        get_args,
    )

from urllib.parse import unquote

from openpyxl import load_workbook
from pydantic import (
    AnyHttpUrl,
    BaseModel,
    FilePath,
    StrictStr,
    ValidationError,
    parse_obj_as,
    validator,
)
from pydantic.fields import ModelField
import requests

from .api import current_session
from .utils import raise_for_status

logger = logging.getLogger(__name__)

FieldId = TypeVar('FieldId', bound=str)


def shape_file_type(path: Path) -> Type[Segmentation] | Type[Mesh]:
    """
    Determine the type of the shape file.

    As described in https://github.com/SCIInstitute/ShapeWorks/blob/3344bbfd42cb83eea50d01c06218e3034b5c67aa/Libs/Mesh/Mesh.h#L212,
    Meshes are of file type "vtk", "vtp", "ply", "stl", "obj", while Segmentations are any other
    file type.
    """  # noqa: E501
    file_type = path.suffix
    if file_type in ['.vtk', '.vtp', '.ply', '.stl', '.obj']:
        return Mesh
    else:
        return Segmentation


class NonEmptyString(StrictStr):
    min_length = 1


class FileType(Generic[FieldId]):
    def __init__(
        self,
        path: Optional[Union[Path, str]] = None,
        url: Optional[str] = None,
        field_id: Optional[str] = None,
    ):
        self.field_id: Optional[str] = field_id
        self.path: Optional[Path] = None
        self.url: Optional[AnyHttpUrl] = None
        self.field_value: Optional[str] = None

        if path:
            self.path = parse_obj_as(FilePath, path)

        if url:
            self.url = parse_obj_as(AnyHttpUrl, url)

    @classmethod
    def __get_validators__(cls):
        yield cls.validate

    @classmethod
    def validate(cls, v, field: ModelField, **kwargs):
        # this is probably a bad idea...
        if not field.sub_fields:
            raise SyntaxError('A field id must be provided when using FileType')

        field_id = get_args(field.sub_fields[0].type_)[0]
        if isinstance(v, FileType):
            v.field_id = field_id
            return v

        path = None
        url = None
        try:
            path = parse_obj_as(FilePath, v)
        except ValidationError:
            pass

        try:
            url = parse_obj_as(AnyHttpUrl, v)
        except ValidationError:
            pass

        if path is None and url is None:
            raise ValueError(f'Could not parse {v} as a local path or a remote url')

        return FileType(path=path, url=url, field_id=field_id)

    def upload(self):
        session = current_session()

        if self.field_value:
            return self.field_value

        if self.url is not None or self.path is None:
            # trying to upload a file when the model was sourced from the server
            raise Exception('Cannot upload a remote file reference')

        if self.field_id is None:
            # I assume validate will always get called, but pydantic *might* be doing something
            # unusual in some cases.
            raise Exception('Unknown field id, this is likely a bug in the FileType class')

        with self.path.open('rb') as f:
            logger.info('Uploading file %s', self.path.name)
            self.field_value = session.s3ff.upload_file(f, str(self.path.name), self.field_id)
            logger.debug('Uploaded file %s', self.path.name)

        return self.field_value

    def download(self, path: Union[Path, str]) -> Path:
        if self.url is None:
            raise Exception('Cannot download a local file')

        if not self.url.path:
            raise Exception('Server returned an unexpected url')

        path = Path(path)
        if path.is_file():
            raise ValueError('Expected a directory name')

        if not path.is_dir():
            path.mkdir(parents=True, exist_ok=True)

        path = path / self.url.path.split('/')[-1]
        r = requests.get(self.url, stream=True)
        raise_for_status(r)

        logger.info('Downloading %s', path)
        with path.open('wb') as f:
            for chunk in r.iter_content(chunk_size=8192):
                f.write(chunk)
        return path

    @property
    def name(self) -> str:
        if self.path:
            return self.path.name
        elif self.url:
            if self.url.path is None:
                raise Exception('Invalid file url')
            return unquote(self.url.path.split('/')[-1])
        raise Exception('Invalid file object')

    def __str__(self):
        return self.name


ModelType = TypeVar('ModelType', bound='ApiModel')


class ApiModel(BaseModel):
    _endpoint: str

    id: Optional[int]

    @validator('*', pre=True)
    def fetch_entity(cls, v, field: ModelField):
        type_ = cls.__fields__[field.name].type_
        if type_ is not Any and issubclass(type_, ApiModel) and isinstance(v, int):
            return type_.from_id(v)
        return v

    @classmethod
    def from_id(cls: Type[ModelType], id: int, **kwargs) -> ModelType:
        session = current_session()
        cache = session.cache[cls]

        if id not in cache:
            r: requests.Response = session.get(f'{cls._endpoint}/{id}/')
            raise_for_status(r)
            json = r.json()
            for key, value in cls.__fields__.items():
                if key in kwargs:
                    json[key] = kwargs[key]
                elif value.type_ is not Any and issubclass(value.type_, ApiModel):
                    json[key] = value.type_.from_id(json[key])
            cache[id] = cls(**json)
        return cache[id]

    @classmethod
    def list(cls: Type[ModelType], **kwargs) -> Iterator[ModelType]:
        session = current_session()

        filter: Dict[str, Any] = {}
        replace: Dict[str, ApiModel] = {}
        for key, value in kwargs.items():
            if isinstance(value, ApiModel):
                filter[key] = value.id
                replace[key] = value
            else:
                filter[key] = value

        r: requests.Response = session.get(f'{cls._endpoint}/', params=filter)

        while True:
            raise_for_status(r)
            data = r.json()
            for result in data['results']:
                result.update(replace)
                yield cls(**result)
            if not data.get('next'):
                return

            r = session.get(data['next'])

    def delete(self) -> None:
        session = current_session()

        self.assert_remote()
        r: requests.Response = session.delete(f'{self._endpoint}/{self.id}/')
        raise_for_status(r)
        self.id = None

    def create(self: ModelType) -> ModelType:
        session = current_session()

        self.assert_local()
        json = self.dict()
        for key, value in self:
            if isinstance(value, FileType):
                json[key] = value.upload()
            if isinstance(value, ApiModel):
                if value.id is None:
                    value.create()
                json[key] = value.id

        r: requests.Response = session.post(f'{self._endpoint}/', json=json)
        raise_for_status(r)
        self.id = r.json()['id']
        return self

    def download_files(self, path: Union[Path, str]) -> Iterator[Path]:
        for _, value in self:
            if isinstance(value, FileType):
                yield value.download(path)

    def assert_remote(self):
        if self.id is None:
            raise Exception('This entity has not yet been created on the server.')

    def assert_local(self):
        if self.id is not None:
            raise Exception('This entity already exists on the server (id: %r).' % self.id)


class Dataset(ApiModel):
    _endpoint = 'datasets'

    name: NonEmptyString
    file: Optional[FileType[Literal['core.Dataset.file']]] = None
    license: NonEmptyString
    description: NonEmptyString
    acknowledgement: NonEmptyString
    keywords: str = ''
    contributors: str = ''
    publications: str = ''

    def __repr_args__(self: BaseModel):
        return [
            (key, value)
            for key, value in self.__dict__.items()
            if key
            not in {
                'license',
                'description',
                'acknowledgement',
                'keywords',
                'contributors',
                'publications',
            }
        ]

    @property
    def subjects(self) -> Iterator[Subject]:
        self.assert_remote()
        return Subject.list(dataset=self)

    def add_subject(self, name: str) -> Subject:
        return Subject(name=name, dataset=self).create()

    @property
    def projects(self) -> Iterator[Project]:
        self.assert_remote()
        return Project.list(dataset=self)

    @property
    def segmentations(self) -> Iterator[Segmentation]:
        for subject in self.subjects:
            for segmentation in subject.segmentations:
                yield segmentation

    @property
    def meshes(self) -> Iterator[Mesh]:
        for subject in self.subjects:
            for mesh in subject.meshes:
                yield mesh

    @property
    def images(self) -> Iterator[Image]:
        for subject in self.subjects:
            for mesh in subject.images:
                yield mesh

    def force_create(self, backup=False):
        """
        Forcibly create the Dataset, even if it already exists.

        If backup=False (the default), then the existing Dataset is deleted.
        If backup=True, then the new Dataset will append an appropriate `-v*` version string to
        its name before creation.
        """
        old_dataset = Dataset.from_name(self.name)
        while old_dataset is not None:
            if not backup:
                # Delete the old dataset to resolve the collision
                old_dataset.delete()
            else:
                # Use a version suffix to resolve the collision
                version_regex = re.compile('(.*)-v([0-9]+)')
                match = version_regex.match(self.name)
                if match:
                    # The old name had a suffix, so increment it
                    name, old_version = match.groups()
                    logger.info('Trying a new name: %s, %s', name, old_version)
                    new_version = int(old_version) + 1
                    self.name = f'{name}-v{new_version}'  # type: ignore
                else:
                    # The old name had no suffix, so append "-v1"
                    self.name = f'{self.name}-v1'  # type: ignore
            # We have a new name now, but that new name might also conflict.
            # Keep looping until there is no conflict.
            old_dataset = Dataset.from_name(self.name)
        return self.create()

    def add_project(self, file: Path, keywords: str = '', description: str = '') -> Project:
        project = Project(
            file=file,
            keywords=keywords,
            description=description,
            dataset=self,
        ).create()
        return project._load_project_spreadsheet()

    @classmethod
    def from_name(cls, name: str) -> Optional[Dataset]:
        results = cls.list(name=name)
        try:
            return next(results)
        except StopIteration:
            return None

    def create(self) -> Dataset:
        result = super().create()
        if self.file:
            self._load_data_spreadsheet()
        # Load the new dataset so we get an appropriate file field
        assert result.id
        return Dataset.from_id(result.id)

    def _load_data_spreadsheet(self):  # noqa: C901
        if not self.file or not self.file.path:
            return
        file = self.file.path

        xls = load_workbook(str(file), read_only=True)
        if 'data' not in xls:
            raise Exception('`data` sheet not found')

        data = xls['data'].values

        headers = next(data)
        for header in headers:
            if not (header.startswith('shape_') or header.startswith('image_')):
                raise Exception(
                    f'Unknown spreadsheet format in {file} - expected "shape_" or "image_" prefix, found {header}'  # noqa: E501
                )
        # split each header into ('shape', anatomy_type) and ('image', modality) tuples
        column_info = [header.split('_', 1) for header in headers]
        # at least one shape column must be present
        if all(info[0] != 'shape' for info in column_info):
            raise Exception('No "shape_" column specified')

        root = file.parent
        subjects: Dict[str, Subject] = {}

        found = False
        for row in data:
            subject = None
            for i, cell in enumerate(row):
                if not cell:
                    continue
                file_type, domain = column_info[i]
                file_path: Path = root / cell

                # Use the file name in the first cell as the subject name
                if subject is None:
                    subject_name = file_path.stem
                    if subject_name not in subjects:
                        subjects[subject_name] = self.add_subject(subject_name)
                    subject = subjects[subject_name]

                if file_type == 'shape':
                    shape_file = file_path
                    if not shape_file.exists():
                        raise Exception(f'Could not find shape file at "{shape_file}"')
                    data_type = shape_file_type(shape_file)
                    if data_type == Segmentation:
                        subject.add_segmentation(file=shape_file, anatomy_type=domain)
                    elif data_type == Mesh:
                        subject.add_mesh(file=shape_file, anatomy_type=domain)
                    found = True
                elif file_type == 'image':
                    image_file = file_path
                    if not image_file.exists():
                        raise Exception(f'Could not find image file at "{image_file}"')
                    subject.add_image(file=image_file, modality=domain)
                    found = True
        if not found:
            raise Exception('Did not find any shape or image files in data sheet')

    def download(self, path: Union[Path, str]):
        self.assert_remote()
        path = Path(path)
        dataset_file = None
        if self.file:
            dataset_file = self.file.download(path)
        for project in self.projects:
            project.download(path)
        if not dataset_file:
            return
        xls = load_workbook(str(dataset_file), read_only=True)
        for subject in self.subjects:
            row = next(entry for entry in xls['data'].values if subject.name == Path(entry[0]).stem)
            for segmentation in subject.segmentations:
                segmentation.file.download(path / Path(row[0]).parent)
            for mesh in subject.meshes:
                mesh.file.download(path / Path(row[0]).parent)
            for image in subject.images:
                image.file.download(path / Path(row[1]).parent)


class Subject(ApiModel):
    _endpoint = 'subjects'

    name: NonEmptyString
    dataset: Dataset

    @property
    def segmentations(self) -> Iterator[Segmentation]:
        self.assert_remote()
        return Segmentation.list(subject=self)

    @property
    def meshes(self) -> Iterator[Mesh]:
        self.assert_remote()
        return Mesh.list(subject=self)

    @property
    def images(self) -> Iterator[Image]:
        self.assert_remote()
        return Image.list(subject=self)

    def add_segmentation(self, file: Path, anatomy_type: str) -> Segmentation:
        return Segmentation(file=file, anatomy_type=anatomy_type, subject=self).create()

    def add_mesh(self, file: Path, anatomy_type: str) -> Mesh:
        return Mesh(file=file, anatomy_type=anatomy_type, subject=self).create()

    def add_image(self, file: Path, modality: str) -> Image:
        return Image(file=file, modality=modality, subject=self).create()


class Segmentation(ApiModel):
    _endpoint = 'segmentations'

    file: FileType[Literal['core.Segmentation.file']]
    anatomy_type: NonEmptyString
    subject: Subject


class Mesh(ApiModel):
    _endpoint = 'meshes'

    file: FileType[Literal['core.Mesh.file']]
    anatomy_type: NonEmptyString
    subject: Subject


class Image(ApiModel):
    _endpoint = 'images'

    file: FileType[Literal['core.Image.file']]
    modality: str
    subject: Subject


class Project(ApiModel):
    _endpoint = 'projects'

    file: FileType[Literal['core.Project.file']]
    keywords: str = ''
    description: str = ''
    dataset: Dataset

    @property
    def groomed_segmentations(self) -> Iterator[GroomedSegmentation]:
        self.assert_remote()
        return GroomedSegmentation.list(project=self)

    @property
    def groomed_meshes(self) -> Iterator[GroomedMesh]:
        self.assert_remote()
        return GroomedMesh.list(project=self)

    def add_groomed_segmentation(
        self,
        file: Path,
        segmentation: Segmentation,
        pre_cropping: Optional[Path] = None,
        pre_alignment: Optional[Path] = None,
    ) -> GroomedSegmentation:
        return GroomedSegmentation(
            file=file,
            segmentation=segmentation,
            project=self,
            pre_cropping=pre_cropping,
            pre_alignment=pre_alignment,
        ).create()

    def add_groomed_mesh(
        self,
        file: Path,
        mesh: Mesh,
        pre_cropping: Optional[Path] = None,
        pre_alignment: Optional[Path] = None,
    ) -> GroomedMesh:
        return GroomedMesh(
            file=file,
            mesh=mesh,
            project=self,
            pre_cropping=pre_cropping,
            pre_alignment=pre_alignment,
        ).create()

    @property
    def shape_models(self) -> Iterator[OptimizedShapeModel]:
        self.assert_remote()
        return OptimizedShapeModel.list(project=self)

    def add_shape_model(self, parameters: Dict[str, Any]) -> OptimizedShapeModel:
        return OptimizedShapeModel(
            project=self,
            parameters=parameters,
        ).create()

    def _load_project_spreadsheet(self) -> Project:
        file = self.file.path
        assert file  # should be guaranteed by assert_local

        root = file.parent
        xls = load_workbook(str(file), read_only=True)

        segmentations = {
            PurePath(segmentation.file.name).stem: segmentation
            for segmentation in self.dataset.segmentations
        }
        meshes = {PurePath(mesh.file.name).stem: mesh for mesh in self.dataset.meshes}

        # if 'optimize' not in xls or 'data' not in xls:
        #     raise Exception('`data` sheet not found')

        # shape_model = self._parse_optimize_sheet(xls['optimize'].values)
        # self._parse_data_sheet(segmentations, shape_model, xls['data'].values, root)

        for sheet_name in ['optimize', 'data']:
            if sheet_name not in xls:
                raise Exception(f'`{sheet_name}` sheet not found')

        shape_model = self._parse_optimize_sheet(xls['optimize'].values)
        self._parse_data_sheet(segmentations, meshes, shape_model, xls['data'], root)

        return self

    def _iter_data_sheet(
        self, sheet: Any, root: Path
    ) -> Iterator[Tuple[Path, Path, str, Path, Path, Path]]:
        # The old style ended in _file.  The newer style ends in an integer
        # (_1, _2, ...).  These can be in different orders
        # ##DWM::
        required = [
            'shape',
            'groomed',
            'alignment',
            'local_particles',
            'world_particles',
        ]
        optional = [
            'procrustes',
            'constraints',
        ]
        headers = next(sheet.values)
        if not any(all(k + suffix in headers for k in required) for suffix in ('_file', '_1')):
            raise Exception(
                'Unknown spreadsheet format - expected headers to include %r '
                'with a suffix of either _file or _1, found %r' % (required, headers)
            )
        for idx in itertools.count():
            suffix = '_file' if not idx else '_%d' % idx
            if not all(k + suffix in headers for k in required):
                if idx >= 1:
                    break
                continue
            col = {
                k: headers.index(k + suffix) for k in required + optional if k + suffix in headers
            }
            rows = sheet.values
            next(rows)  # skip header
            for rowrec in rows:
                row = {k: rowrec[col[k]] for k in col}
                if not row['shape']:
                    # It's possible to get XLSX files with empty rows where every column is None
                    continue

                shape_file = root / row['shape']
                groomed_file = root / row['groomed']
                local = root / row['local_particles']
                world = root / row['world_particles']
                alignment_file = row['alignment']
                constraints = (root / row['constraints']) if 'constraints' in row else None

                yield shape_file, groomed_file, alignment_file, local, world, constraints

    def _parse_optimize_sheet(self, sheet: Any) -> OptimizedShapeModel:
        expected = ('key', 'value')
        headers = next(sheet)
        if headers[: len(expected)] != expected:
            raise Exception(
                'Unknown spreadsheet format - expected headers to be %r, found %r'
                % (expected, headers[: len(expected)])
            )

        params: Dict[str, Union[str, float]] = {}
        for row in sheet:
            key, value = row
            try:
                value = float(value)
            except ValueError:
                pass
            params[key] = value

        return self.add_shape_model(params)

    def _parse_data_sheet(
        self,
        segmentations: Dict[str, Segmentation],
        meshes: Dict[str, Mesh],
        shape_model: OptimizedShapeModel,
        sheet: Any,
        root: Path,
    ):
        for (
            shape_file,
            groomed_file,
            alignment_file,
            local,
            world,
            constraints,
        ) in self._iter_data_sheet(sheet, root):
            groomed_segmentation = None
            groomed_mesh = None
            data_type = shape_file_type(shape_file)
            if data_type == Segmentation:
                segmentation = segmentations.get(shape_file.stem)
                if not segmentation:
                    raise Exception(f'Could not find segmentation for "{shape_file}"')
                groomed_segmentation = self.add_groomed_segmentation(
                    file=groomed_file,
                    segmentation=segmentation,
                )
            elif data_type == Mesh:
                mesh = meshes.get(shape_file.stem)
                if not mesh:
                    raise Exception(f'Could not find mesh for "{shape_file}"')
                groomed_mesh = self.add_groomed_mesh(
                    file=groomed_file,
                    mesh=mesh,
                )

            with TemporaryDirectory() as dir:
                transform = Path(dir) / f'{shape_file.stem}.transform'
                with transform.open('w') as f:
                    f.write(alignment_file)

                shape_model.add_particles(
                    world=world,
                    local=local,
                    transform=transform,
                    groomed_segmentation=groomed_segmentation,
                    groomed_mesh=groomed_mesh,
                    constraints=constraints,
                )

    def download(self, path: Union[Path, str]):
        path = Path(path)
        project_file = self.file.download(path)
        xls = load_workbook(str(project_file), read_only=True)
        sheet = xls['data']

        shape_model = next(self.shape_models)
        groomed_segmentations = {
            PurePath(gs.file.name).stem: gs for gs in self.groomed_segmentations
        }
        local_files = {PurePath(p.local.name).stem: p for p in shape_model.particles}

        # TODO: Do we detect if alignment_file (transform) is a path?
        for _, groomed_file, _, local, world, constraints in self._iter_data_sheet(sheet, path):
            gs = groomed_segmentations[groomed_file.stem]
            gs.file.download(groomed_file.parent)

            particles = local_files[local.stem]
            particles.local.download(local.parent)
            particles.world.download(world.parent)
            if particles.constraints and constraints:
                particles.constraints.download(constraints.parent)


class GroomedSegmentation(ApiModel):
    _endpoint = 'groomed-segmentations'

    file: FileType[Literal['core.GroomedSegmentation.file']]
    pre_cropping: Optional[FileType[Literal['core.GroomedSegmentation.pre_cropping']]] = None
    pre_alignment: Optional[FileType[Literal['core.GroomedSegmentation.pre_alignment']]] = None

    segmentation: Segmentation
    project: Project


class GroomedMesh(ApiModel):
    _endpoint = 'groomed-meshes'

    file: FileType[Literal['core.GroomedMesh.file']]
    pre_cropping: Optional[FileType[Literal['core.GroomedMesh.pre_cropping']]] = None
    pre_alignment: Optional[FileType[Literal['core.GroomedMesh.pre_alignment']]] = None

    mesh: Mesh
    project: Project


class OptimizedShapeModel(ApiModel):
    _endpoint = 'optimized-shape-models'

    project: Project
    parameters: Dict[str, Any]

    @property
    def particles(self) -> Iterator[OptimizedParticles]:
        return OptimizedParticles.list(shape_model=self)

    def add_particles(
        self,
        world: Path,
        local: Path,
        transform: Path,
        groomed_segmentation: Optional[GroomedSegmentation],
        groomed_mesh: Optional[GroomedMesh],
        constraints: Path,
    ) -> OptimizedParticles:
        return OptimizedParticles(
            world=world,
            local=local,
            transform=transform,
            shape_model=self,
            groomed_segmentation=groomed_segmentation,
            groomed_mesh=groomed_mesh,
            constraints=constraints,
        ).create()


class OptimizedParticles(ApiModel):
    _endpoint = 'optimized-particles'

    world: FileType[Literal['core.OptimizedParticles.world']]
    local: FileType[Literal['core.OptimizedParticles.local']]
    transform: FileType[Literal['core.OptimizedParticles.transform']]
    shape_model: OptimizedShapeModel
    groomed_segmentation: Optional[GroomedSegmentation]
    groomed_mesh: Optional[GroomedMesh]
    constraints: Optional[FileType[Literal['core.OptimizedParticles.constraints']]] = None


class OptimizedPCAModel(ApiModel):
    _endpoint = 'optimized-pca-model'

    mean_particles: FileType[Literal['core.OptimizedPCAModel.mean_particles']]
    pca_modes: FileType[Literal['core.OptimizedPCAModel.pca_modes']]
    eigen_spectrum: FileType[Literal['core.OptimizedPCAModel.eigen_spectrum']]
    shape_model: OptimizedShapeModel
